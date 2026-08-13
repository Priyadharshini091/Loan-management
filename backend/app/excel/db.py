import os
import shutil
import threading
from datetime import datetime
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl import Workbook
from app.config import settings

excel_lock = threading.Lock()

SHEET_COLUMNS = {
    "Users": [
        "user_id", "username", "email", "password_hash", "role", "status", "created_at"
    ],
    "Areas": [
        "area_id", "area_name", "district", "pincode", "status", "created_at"
    ],
    "Customers": [
        "customer_id", "customer_name", "mobile_number", "address", "area_id", "area_name",
        "aadhaar_number", "photo_path", "guarantor_name", "guarantor_mobile", "created_at", "status"
    ],
    "Loans": [
        "loan_id", "loan_number", "customer_id", "customer_name", "area_id", "area_name",
        "loan_amount", "interest_percentage", "interest_amount", "total_payable", "loan_date",
        "emi_type", "number_of_installments", "emi_amount", "first_due_date", "final_due_date", "status"
    ],
    "Payments": [
        "payment_id", "receipt_number", "loan_id", "customer_id", "customer_name", "area_id", "area_name",
        "payment_date", "due_date", "amount_due", "amount_paid", "partial_payment", "balance",
        "payment_status", "payment_method", "collected_by", "remarks"
    ],
    "Installments": [
        "installment_id", "loan_id", "customer_id", "installment_number", "due_date",
        "due_amount", "paid_amount", "balance", "status", "paid_date", "receipt_number"
    ],
    "Audit_Log": [
        "log_id", "user_id", "username", "action", "module", "record_id", "timestamp", "details"
    ]
}

def init_excel():
    with excel_lock:
        file_path = settings.EXCEL_FILE_PATH
        if not os.path.exists(file_path):
            wb = Workbook()
            # remove default sheet
            default_sheet = wb.active
            for sheet_name, cols in SHEET_COLUMNS.items():
                ws = wb.create_sheet(title=sheet_name)
                ws.append(cols)
            if default_sheet and "Sheet" in default_sheet.title:
                wb.remove(default_sheet)
            wb.save(file_path)
            wb.close()
        else:
            wb = openpyxl.load_workbook(file_path)
            modified = False
            for sheet_name, cols in SHEET_COLUMNS.items():
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(title=sheet_name)
                    ws.append(cols)
                    modified = True
            if modified:
                wb.save(file_path)
            wb.close()

def create_backup() -> str:
    """Create timestamped backup of loan_management.xlsx"""
    with excel_lock:
        file_path = settings.EXCEL_FILE_PATH
        if not os.path.exists(file_path):
            return ""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"loan_management_{timestamp}.xlsx"
        backup_path = os.path.join(settings.BACKUP_DIR, backup_filename)
        shutil.copy2(file_path, backup_path)
        return backup_path

def read_sheet(sheet_name: str) -> List[Dict[str, Any]]:
    init_excel()
    with excel_lock:
        file_path = settings.EXCEL_FILE_PATH
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows or len(rows) < 1:
            return []

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            if not any(row):
                continue
            row_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    val = row[idx]
                    row_dict[header] = "" if val is None else str(val)
                else:
                    row_dict[header] = ""
            result.append(row_dict)
        return result

def write_sheet(sheet_name: str, records: List[Dict[str, Any]], auto_backup: bool = True):
    if auto_backup:
        create_backup()
    init_excel()
    cols = SHEET_COLUMNS.get(sheet_name, [])
    with excel_lock:
        file_path = settings.EXCEL_FILE_PATH
        wb = openpyxl.load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(cols)

        for record in records:
            row = [str(record.get(col, "")) for col in cols]
            ws.append(row)

        wb.save(file_path)
        wb.close()

def log_audit(user_id: str, username: str, action: str, module: str, record_id: str, details: str):
    logs = read_sheet("Audit_Log")
    log_id = f"LOG{len(logs) + 1:06d}"
    new_log = {
        "log_id": log_id,
        "user_id": user_id,
        "username": username,
        "action": action,
        "module": module,
        "record_id": record_id,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "details": details
    }
    logs.append(new_log)
    write_sheet("Audit_Log", logs, auto_backup=False)
